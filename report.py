"""
LedgerBridge AI — Report Generator

Builds the final Excel reconciliation report with:
- Summary sheet (balance walk + counts + RECONCILED banner)
- Matched sheet (all L1/L2/L3 pairs)
- Amount Mismatches sheet
- Missing in Ours sheet
- Missing in Theirs sheet
- Timing Differences sheet (subset of matched, L2 only)
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reconcile import ReconciliationResult


# Colors (RGB hex)
HEADER_FILL = "1F4E78"      # Dark blue
HEADER_FONT = "FFFFFF"      # White
GOOD_FILL = "C6EFCE"        # Light green
GOOD_FONT = "006100"
BAD_FILL = "FFC7CE"         # Light red
BAD_FONT = "9C0006"
WARN_FILL = "FFEB9C"        # Light yellow
WARN_FONT = "9C5700"
BORDER_COLOR = "B0B0B0"

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)


def write_report(
    result: ReconciliationResult,
    output_path: str | Path,
    ai_insights: str = "",
    tds_result=None,
    cost_tracker=None,
) -> Path:
    """
    Generate the full Excel report at output_path.

    Args:
        result:       The reconciliation result from reconcile.py
        output_path:  Where to save the .xlsx
        ai_insights:  Optional plain-text AI insights paragraph
        tds_result:   Optional TdsReconciliationResult — adds "TDS Reconciliation" sheet
        cost_tracker: Optional CostTracker — adds API cost line to Summary

    Returns the path written.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # Drop default sheet

    _write_summary(wb, result, ai_insights, tds_result=tds_result, cost_tracker=cost_tracker)
    if tds_result is not None:
        _write_tds_reconciliation(wb, tds_result)
    _write_matched(wb, result)
    _write_needs_review(wb, result)
    _write_amount_mismatches(wb, result)
    _write_missing_theirs(wb, result, tds_result=tds_result)
    _write_missing_ours(wb, result, tds_result=tds_result)
    _write_timing(wb, result)

    wb.save(output_path)
    return output_path


def write_standardized_ledger(df: pd.DataFrame, output_path: str | Path, title: str) -> Path:
    """Write a single standardized ledger as a formatted Excel file."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Standardized"

    # Title row
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    _write_df_to_sheet(ws, df, start_row=3)
    _auto_column_widths(ws, df, start_row=3)

    wb.save(output_path)
    return output_path


# ---------------- Individual sheet writers ----------------

def _write_summary(wb: Workbook, result: ReconciliationResult, ai_insights: str, tds_result=None, cost_tracker=None) -> None:
    ws = wb.create_sheet("Summary")
    s = result.summary

    # Title
    ws["A1"] = "Reconciliation Summary"
    ws["A1"].font = Font(bold=True, size=16, color=HEADER_FILL)
    ws.merge_cells("A1:D1")

    # Match counts
    row = 3
    ws.cell(row=row, column=1, value="Match Statistics").font = Font(bold=True, size=12)
    row += 1

    # Adjust missing counts to exclude rows reclassified as TDS entries
    tds_flagged_ours = len(tds_result.removed_from_missing_theirs) if tds_result else 0
    tds_flagged_theirs = len(tds_result.removed_from_missing_ours) if tds_result else 0
    adj_missing_theirs = s["missing_in_theirs"] - tds_flagged_ours
    adj_missing_ours   = s["missing_in_ours"]   - tds_flagged_theirs

    stats = [
        ("Total records (Our books)",   s["total_our_records"]),
        ("Total records (Their books)", s["total_their_records"]),
        ("L1 matches (Date + Ref + Amount)", s["matched_l1"]),
        ("L2 matches (Ref + Amount, dates differ)", s["matched_l2"]),
        ("L3 matches (Date + Amount, weak)", s["matched_l3"]),
        ("Amount mismatches", s["amount_mismatches"]),
        ("Variance (e.g. bank charge / short payment)", s.get("variance", 0)),
        ("Sign-reversed (possible posting error)", s.get("sign_reversed", 0)),
        ("Suspected duplicates", s.get("suspected_duplicates", 0)),
        ("Missing in their books", adj_missing_theirs),
        ("Missing in our books", adj_missing_ours),
    ]
    if tds_result is not None:
        total_tds_flagged = tds_flagged_ours + tds_flagged_theirs
        if total_tds_flagged > 0:
            stats.append(("TDS journal entries (see TDS Reconciliation sheet)", total_tds_flagged))

    for label, val in stats:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=val)
        c.alignment = Alignment(horizontal="right")
        c.font = Font(bold=True)
        row += 1

    # Balance walk
    row += 2
    ws.cell(row=row, column=1, value="Closing Balance Walk").font = Font(bold=True, size=12)
    row += 1
    walk = [
        ("Opening Balance (Ours)",          s["opening_balance_ours"]),
        ("+ Sum of our transactions",       s["sum_our_transactions"]),
        ("= Closing Balance (Ours)",        s["closing_balance_ours"]),
        ("",                                 None),
        ("Opening Balance (Theirs)",         s["opening_balance_theirs"]),
        ("+ Sum of their transactions",     s["sum_their_transactions"]),
        ("= Closing Balance (Theirs)",      s["closing_balance_theirs"]),
        ("",                                 None),
        ("Net Position (Ours + Theirs, should be ~0)", s["difference"]),
        ("Reconciling items (missing + mismatches)",   s["reconciling_item"]),
        ("Residual (should be 0)",                     s["residual"]),
    ]
    for label, val in walk:
        ws.cell(row=row, column=1, value=label)
        if val is not None:
            c = ws.cell(row=row, column=2, value=val)
            c.number_format = '#,##0.00;(#,##0.00);-'
            c.alignment = Alignment(horizontal="right")
            if "=" in label or "Residual" in label:
                c.font = Font(bold=True)
        row += 1

    # TDS reference
    # When tds_result is available, show the full picture (column totals +
    # journal entries + net gap + plain-English status). Otherwise fall back
    # to the legacy column-only view.
    row += 2
    ws.cell(row=row, column=1, value="TDS Reference").font = Font(bold=True, size=12)
    row += 1

    if tds_result is not None:
        # Compute net gap that aligns with the dedicated TDS Reconciliation sheet:
        #   gap = (our column + our journal) - (their column + their journal)
        our_total   = tds_result.our_tds_column_total   + tds_result.our_tds_journal_total
        their_total = tds_result.their_tds_column_total + tds_result.their_tds_journal_total
        net_gap = our_total - their_total

        tds_rows = [
            ("TDS in our books (column total)",     tds_result.our_tds_column_total),
            ("TDS in their books (column total)",   tds_result.their_tds_column_total),
            ("TDS booked as journal (ours)",        tds_result.our_tds_journal_total),
            ("TDS booked as journal (theirs)",      tds_result.their_tds_journal_total),
            ("Net TDS gap",                          net_gap),
        ]
        for label, val in tds_rows:
            ws.cell(row=row, column=1, value=label)
            c = ws.cell(row=row, column=2, value=val)
            c.number_format = '#,##0.00;(#,##0.00);-'
            c.alignment = Alignment(horizontal="right")
            if label == "Net TDS gap":
                c.font = Font(bold=True)
                ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

        # Status line — colour-coded like the TDS sheet's banner
        row += 1
        status = tds_result.overall_status
        msg    = tds_result.status_message
        if status == "MATCHED":
            fill, font_col, prefix = GOOD_FILL, GOOD_FONT, "✓ MATCHED"
        elif status == "PARTIAL":
            fill, font_col, prefix = WARN_FILL, WARN_FONT, "⚠ PARTIAL"
        elif status == "EXCESS":
            fill, font_col, prefix = BAD_FILL, BAD_FONT, "✗ EXCESS"
        elif status == "NO_TDS_ACTIVITY":
            fill, font_col, prefix = "EEEEEE", "555555", "— NO TDS"
        else:  # UNVERIFIED
            fill, font_col, prefix = WARN_FILL, WARN_FONT, "? UNVERIFIED"

        full_msg = f"Status: {prefix} — {msg}  See TDS Reconciliation sheet for individual entries."
        c = ws.cell(row=row, column=1, value=full_msg)
        c.font = Font(bold=True, color=font_col)
        c.fill = PatternFill("solid", start_color=fill)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 45
        row += 1
    else:
        # Fallback: legacy column-only view (no TDS post-processing was run)
        tds_rows = [
            ("TDS in our books",       s["tds_ours"]),
            ("TDS in their books",     s["tds_theirs"]),
            ("TDS difference",         s["tds_difference"]),
        ]
        for label, val in tds_rows:
            ws.cell(row=row, column=1, value=label)
            c = ws.cell(row=row, column=2, value=val)
            c.number_format = '#,##0.00;(#,##0.00);-'
            c.alignment = Alignment(horizontal="right")
            row += 1

    # Banner
    row += 2
    if s["reconciled"]:
        banner = ws.cell(row=row, column=1, value=f"✓ RECONCILED  (residual within tolerance of ₹{s['amount_tolerance']:.2f})")
        banner.font = Font(bold=True, size=14, color=GOOD_FONT)
        banner.fill = PatternFill("solid", start_color=GOOD_FILL)
    else:
        banner = ws.cell(row=row, column=1, value=f"✗ NOT RECONCILED  (residual: ₹{s['residual']:,.2f})")
        banner.font = Font(bold=True, size=14, color=BAD_FONT)
        banner.fill = PatternFill("solid", start_color=BAD_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    # API cost (if tracker was provided)
    if cost_tracker is not None and cost_tracker.records:
        totals = cost_tracker.total()
        row += 3
        ws.cell(row=row, column=1, value="API Cost (this reconciliation)").font = Font(bold=True, size=12)
        row += 1
        cost_rows = [
            ("Total input tokens",   totals["input_tokens"]),
            ("Total output tokens",  totals["output_tokens"]),
            ("API calls made",       totals["calls"]),
            ("Cached steps (skipped API)", totals["cache_hits"]),
            ("Cost (USD)",           f"${totals['cost_usd']:.4f}"),
            ("Cost (INR, approx)",   f"₹{totals['cost_inr']:.2f}"),
        ]
        for label, val in cost_rows:
            ws.cell(row=row, column=1, value=label)
            c = ws.cell(row=row, column=2, value=val)
            c.alignment = Alignment(horizontal="right")
            if "Cost" in label:
                c.font = Font(bold=True)
            row += 1

    # AI insights
    if ai_insights:
        row += 3
        ws.cell(row=row, column=1, value="AI Insights").font = Font(bold=True, size=12)
        row += 1
        c = ws.cell(row=row, column=1, value=ai_insights)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 8, end_column=4)
        ws.row_dimensions[row].height = 200

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22


def _write_tds_reconciliation(wb: Workbook, tds_result) -> None:
    """Write the TDS Reconciliation sheet (only if tds_result is provided)."""
    ws = wb.create_sheet("TDS Reconciliation")

    # ── Title ────────────────────────────────────────────────────────────────
    ws["A1"] = "TDS Reconciliation"
    ws["A1"].font = Font(bold=True, size=16, color=HEADER_FILL)
    ws.merge_cells("A1:E1")

    # ── Status banner ────────────────────────────────────────────────────────
    row = 3
    status = tds_result.overall_status
    status_msg = tds_result.status_message

    if status == "MATCHED":
        fill, font_col, prefix = GOOD_FILL, GOOD_FONT, "✓ MATCHED"
    elif status == "PARTIAL":
        fill, font_col, prefix = WARN_FILL, WARN_FONT, "⚠ PARTIAL"
    elif status == "EXCESS":
        fill, font_col, prefix = BAD_FILL, BAD_FONT, "✗ EXCESS"
    elif status == "NO_TDS_ACTIVITY":
        fill, font_col, prefix = "EEEEEE", "555555", "— NO TDS"
    else:  # UNVERIFIED
        fill, font_col, prefix = WARN_FILL, WARN_FONT, "? UNVERIFIED"

    banner = ws.cell(row=row, column=1, value=f"{prefix}  —  {status_msg}")
    banner.font = Font(bold=True, size=12, color=font_col)
    banner.fill = PatternFill("solid", start_color=fill)
    banner.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 35

    # ── Totals comparison ────────────────────────────────────────────────────
    row += 3
    ws.cell(row=row, column=1, value="TDS Totals Comparison").font = Font(bold=True, size=12)
    row += 1

    totals = [
        ("",                                       "Our Records",                                 "Their Records"),
        ("TDS column total (sum of TDS Amount):",  tds_result.our_tds_column_total,               tds_result.their_tds_column_total),
        ("TDS journal entries (descr-flagged):",   tds_result.our_tds_journal_total,              tds_result.their_tds_journal_total),
    ]
    for i, (label, ours, theirs) in enumerate(totals):
        ws.cell(row=row, column=1, value=label).font = Font(bold=(i == 0))
        if i == 0:
            ws.cell(row=row, column=2, value=ours).font = Font(bold=True)
            ws.cell(row=row, column=3, value=theirs).font = Font(bold=True)
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")
        else:
            for col, val in [(2, ours), (3, theirs)]:
                c = ws.cell(row=row, column=col, value=val)
                c.number_format = '#,##0.00;(#,##0.00);-'
                c.alignment = Alignment(horizontal="right")
        row += 1

    # ── Cross-comparison ─────────────────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="Cross-Comparison").font = Font(bold=True, size=12)
    row += 1

    cross = [
        ("Our TDS column  vs  Their TDS journal:", tds_result.our_tds_column_total - tds_result.their_tds_journal_total),
        ("Their TDS column  vs  Our TDS journal:", tds_result.their_tds_column_total - tds_result.our_tds_journal_total),
    ]
    for label, diff in cross:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=diff)
        c.number_format = '#,##0.00;(#,##0.00);-'
        c.alignment = Alignment(horizontal="right")
        c.font = Font(bold=True)
        row += 1

    # ── Individual flagged entries ───────────────────────────────────────────
    row += 2
    ws.cell(row=row, column=1, value="Individual TDS Journal Entries Detected").font = Font(bold=True, size=12)
    row += 1

    df = tds_result.flagged_entries
    if df.empty:
        ws.cell(row=row, column=1, value="(No TDS-described journal entries detected. "
                                          "TDS is tracked entirely in the TDS column on both sides.)").font = Font(italic=True)
    else:
        # Write as a mini-table
        for col_idx, col_name in enumerate(df.columns, start=1):
            c = ws.cell(row=row, column=col_idx, value=col_name)
            c.font = Font(bold=True, color=HEADER_FONT)
            c.fill = PatternFill("solid", start_color=HEADER_FILL)
            c.alignment = Alignment(horizontal="center")
            c.border = THIN_BORDER
        row += 1
        for _, r in df.iterrows():
            for col_idx, col_name in enumerate(df.columns, start=1):
                v = r[col_name]
                if isinstance(v, pd.Timestamp):
                    v = v.strftime("%Y-%m-%d") if not pd.isna(v) else ""
                elif pd.isna(v):
                    v = ""
                c = ws.cell(row=row, column=col_idx, value=v)
                c.border = THIN_BORDER
                if col_name == "Amount":
                    c.number_format = '#,##0.00;(#,##0.00);-'
                    c.alignment = Alignment(horizontal="right")
                elif col_name == "Status":
                    status_val = str(v)
                    if status_val == "MATCHED":
                        c.fill = PatternFill("solid", start_color=GOOD_FILL)
                        c.font = Font(bold=True, color=GOOD_FONT)
                    elif status_val == "PARTIAL":
                        c.fill = PatternFill("solid", start_color=WARN_FILL)
                        c.font = Font(bold=True, color=WARN_FONT)
                    elif status_val == "EXCESS":
                        c.fill = PatternFill("solid", start_color=BAD_FILL)
                        c.font = Font(bold=True, color=BAD_FONT)
            row += 1

    # ── Column widths ────────────────────────────────────────────────────────
    widths = {"A": 50, "B": 22, "C": 22, "D": 55, "E": 18, "F": 14, "G": 55}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _write_matched(wb: Workbook, result: ReconciliationResult) -> None:
    ws = wb.create_sheet("Matched")
    df = result.matched
    if df.empty:
        ws["A1"] = "(No matched records)"
        return
    _write_df_to_sheet(ws, df, start_row=1)
    _auto_column_widths(ws, df, start_row=1)
    _highlight_amount_cols(ws, df, start_row=1)


def _write_needs_review(wb: Workbook, result: ReconciliationResult) -> None:
    ws = wb.create_sheet("Needs Review")
    df = getattr(result, "needs_review", None)
    if df is None or df.empty:
        ws["A1"] = "(Nothing needs review — every row is a clean match or genuinely missing)"
        return
    # Reason-first table so the accountant can act row-by-row. "Gap" is
    # Ours + Theirs (mirror-sign); ≈0 would be a perfect match.
    _write_df_to_sheet(ws, df, start_row=1, header_fill=WARN_FILL)
    _auto_column_widths(ws, df, start_row=1)
    _highlight_amount_cols(ws, df, start_row=1)


def _write_amount_mismatches(wb: Workbook, result: ReconciliationResult) -> None:
    ws = wb.create_sheet("Amount Mismatches")
    df = result.amount_mismatches
    if df.empty:
        ws["A1"] = "(No amount mismatches — clean!)"
        return
    _write_df_to_sheet(ws, df, start_row=1, header_fill=WARN_FILL)
    _auto_column_widths(ws, df, start_row=1)
    _highlight_amount_cols(ws, df, start_row=1)


def _write_missing_theirs(wb: Workbook, result: ReconciliationResult, tds_result=None) -> None:
    ws = wb.create_sheet("Missing in Their Books")
    df = result.missing_in_theirs
    # Drop rows that have been reclassified as TDS entries
    if tds_result is not None and tds_result.removed_from_missing_theirs:
        df = df.drop(index=tds_result.removed_from_missing_theirs, errors="ignore").reset_index(drop=True)
    if df.empty:
        ws["A1"] = "(Nothing missing on their side)"
        return
    _write_df_to_sheet(ws, df, start_row=1)
    _auto_column_widths(ws, df, start_row=1)
    _highlight_amount_cols(ws, df, start_row=1)


def _write_missing_ours(wb: Workbook, result: ReconciliationResult, tds_result=None) -> None:
    ws = wb.create_sheet("Missing in Our Books")
    df = result.missing_in_ours
    # Drop rows that have been reclassified as TDS entries
    if tds_result is not None and tds_result.removed_from_missing_ours:
        df = df.drop(index=tds_result.removed_from_missing_ours, errors="ignore").reset_index(drop=True)
    if df.empty:
        ws["A1"] = "(Nothing missing on our side)"
        return
    _write_df_to_sheet(ws, df, start_row=1)
    _auto_column_widths(ws, df, start_row=1)
    _highlight_amount_cols(ws, df, start_row=1)


def _write_timing(wb: Workbook, result: ReconciliationResult) -> None:
    ws = wb.create_sheet("Timing Differences")
    df = result.timing_differences
    if df.empty:
        ws["A1"] = "(No timing differences detected)"
        return
    _write_df_to_sheet(ws, df, start_row=1)
    _auto_column_widths(ws, df, start_row=1)
    _highlight_amount_cols(ws, df, start_row=1)


# ---------------- Helpers ----------------

def _write_df_to_sheet(ws, df: pd.DataFrame, start_row: int, header_fill: str = HEADER_FILL) -> None:
    """Write a DataFrame to a worksheet with header styling and borders."""
    # Header row
    for col_idx, col_name in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=col_idx, value=str(col_name))
        c.font = Font(bold=True, color=HEADER_FONT if header_fill == HEADER_FILL else "000000")
        c.fill = PatternFill("solid", start_color=header_fill)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for col_idx, col_name in enumerate(df.columns, start=1):
            value = row[col_name]
            # Convert pandas Timestamps to date strings
            if isinstance(value, pd.Timestamp):
                value = value.strftime("%Y-%m-%d") if not pd.isna(value) else ""
            elif pd.isna(value):
                value = ""
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")

    # Freeze the header row
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def _auto_column_widths(ws, df: pd.DataFrame, start_row: int) -> None:
    """Set sensible column widths based on max content length."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        for value in df[col_name].head(200):  # cap scan to first 200 rows for speed
            v_str = str(value) if not pd.isna(value) else ""
            if len(v_str) > max_len:
                max_len = len(v_str)
        width = min(max(max_len + 2, 10), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _highlight_amount_cols(ws, df: pd.DataFrame, start_row: int) -> None:
    """Apply number formatting to columns whose name contains 'Amount' or 'Difference'."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        n = str(col_name).lower()
        if any(k in n for k in ("amount", "difference", "balance", "total")):
            for r in range(start_row + 1, start_row + 1 + len(df)):
                ws.cell(row=r, column=col_idx).number_format = '#,##0.00;(#,##0.00);-'
                ws.cell(row=r, column=col_idx).alignment = Alignment(horizontal="right")